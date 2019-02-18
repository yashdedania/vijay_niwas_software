from PyQt4.QtCore import *
from PyQt4.QtGui import *
from PyQt4 import QtCore
from PyQt4.QtWebKit import QWebView
from openpyxl import load_workbook
import datetime, sys
import tablib
from dateutil import parser
from jinja2 import Environment, PackageLoader
import ctypes
import warnings

warnings.filterwarnings("ignore")

myappid = 'vijay_niwas.1.0' # arbitrary string
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

font_size = 11
font_style = "Monospace"
excel_file_path = "database/vijay_niwas.xlsx"
icon_path = "database/'vijay_niwas.png"
pdf_template_path = "database/pdf_templates/"
receipt_path = "database/receipts/"
entry_val = ('ENTRY NO','DATE ARRIVAL','NAME','AGE','NATIONALITY','ADDRESS','ARRIVED FROM','ADDRESS TO WHICH PRECEDING',
	'UNIQUE ID','PURPOSE OF VISIT','OCCUPATION','PERSON COUNT','ROOM NO','PRICE','DATE DEPARTURE','REMARKS','CONTACT NO')

tentryno=[]
tdatearr=[]
tname=[]
tpcnt=[]
tdatedeparture=[]
troom=[]
trowno=[]

def main(args):

	class MyTable(QTableWidget):
	    def __init__(self, data, *args):
	        QTableWidget.__init__(self, *args)
	        self.data = data
	        self.setmydata2(args)
	        self.resizeColumnsToContents()
	        header = self.horizontalHeader()
	        header.setStretchLastSection(True)

	    def setmydata2(self, args):
	        horHeaders = ["ENTRY NO","NAME","ROOM NO","PERSON COUNT","DATE ARRIVAl","DATE DEPARTURE"]
	        n = 0
	        for i in range(len(self.data[0])):
	        	rowPosition = self.rowCount()
	        	self.insertRow(rowPosition)
	        	for j in range(args[1]):
	        		newitem = QTableWidgetItem(self.data[j][i])
	        		newitem.setFlags(QtCore.Qt.ItemIsEnabled)
	        		self.setItem(i, j, newitem)
	        self.setHorizontalHeaderLabels(horHeaders)

	
	def  cellClick(trow,tcol):
		global tentryno, trowno
		entry_no = tentryno[trow]
		row_no = trowno[trow]
		menu = QMenu()
		view_det = menu.addAction('View / Edit Details')
		depart = menu.addAction('Departure')
		delete = menu.addAction('Delete Entry')
		action = menu.exec_(QCursor.pos())

		if action == view_det:
			wb = load_workbook(excel_file_path)
			ws = wb.active
			row_cnt = ws.max_row
			column_cnt = ws.max_column
			for index,row in enumerate(ws.iter_rows()):
				if index>0 and index<row_cnt and str(row[0].value)==str(entry_no):
					date_arr = str(row[1].value)
					name = str(row[2].value)
					age = str(row[3].value)
					nationality = str(row[4].value)
					addr = str(row[5].value)
					arrived_frm = str(row[6].value)
					addr_to = str(row[7].value)
					unique_id = str(row[8].value)
					purpose = str(row[9].value)
					occupation = str(row[10].value)
					person_cnt = str(row[11].value)
					date_dept = str(row[14].value)
					remarks = str(row[15].value)
					room_no = str(row[12].value)
					price = str(row[13].value)
					contact_no = str(row[16].value)
					break

			popup = QDialog()
			form = QFormLayout()

			e46 = QLineEdit(entry_no)
			e46.setFont(QFont(font_style,font_size))
			e46.setValidator(QIntValidator())
			form.addRow("ENTRY NO",e46)
			e30 = QLineEdit(name)
			e30.setFont(QFont(font_style,font_size))
			form.addRow("NAME",e30)
			e45 = QLineEdit(contact_no)
			e45.setFont(QFont(font_style,font_size))
			e45.setValidator(QIntValidator())
			form.addRow("CONTACT NO",e45)			
			e31 = QLineEdit(date_arr)
			e31.setReadOnly(True)
			e31.setFont(QFont(font_style,font_size))
			form.addRow("DATE ARRIVAL",e31)
			e32 = QLineEdit(age)
			e32.setFont(QFont(font_style,font_size))
			e32.setValidator(QIntValidator())
			e32.setMaxLength(3)
			form.addRow("AGE",e32)
			e33 = QLineEdit(nationality)
			e33.setFont(QFont(font_style,font_size))
			form.addRow("NATIONALITY",e33)
			e34 = QLineEdit(addr)
			e34.setFont(QFont(font_style,font_size))
			form.addRow("ADDRESS",e34)
			e35 = QLineEdit(arrived_frm)
			e35.setFont(QFont(font_style,font_size))
			form.addRow("ARRIVED FROM",e35)			
			e36 = QLineEdit(addr_to)
			e36.setFont(QFont(font_style,font_size))
			form.addRow("ADDRESS TO PRECEDING",e36)
			e37 = QLineEdit(unique_id)
			e37.setFont(QFont(font_style,font_size))
			form.addRow("UNIQUE ID",e37)			
			e38 = QLineEdit(purpose)
			e38.setFont(QFont(font_style,font_size))
			form.addRow("PURPOSE",e38)
			e39 = QLineEdit(occupation)
			e39.setFont(QFont(font_style,font_size))
			form.addRow("OCCUPATION",e39)
			e40 = QLineEdit(person_cnt)
			e40.setFont(QFont(font_style,font_size))
			e40.setValidator(QIntValidator())
			e40.setMaxLength(3)			
			form.addRow("PERSON COUNT",e40)
			e41 = QLineEdit(room_no)
			e41.setFont(QFont(font_style,font_size))
			e41.setValidator(QIntValidator())
			form.addRow("ROOM NO",e41)
			e42 = QLineEdit(price)
			e42.setFont(QFont(font_style,font_size))
			e42.setValidator(QIntValidator())		
			form.addRow("PRICE",e42)						
			e43 = QLineEdit(date_dept)
			e43.setReadOnly(True)
			e43.setFont(QFont(font_style,font_size))
			form.addRow("DATE DEPARTURE",e43)
			e44 = QLineEdit(remarks)
			e44.setFont(QFont(font_style,font_size))
			form.addRow("REMARKS",e44)
			update_btn = QPushButton("UPDATE")
			update_btn.setFont(QFont("Courier",font_size))	
			form.addRow(update_btn)
			update_btn.clicked.connect(lambda:update_btn_clkd(row_no,e30.text(),e32.text(),e33.text(),e34.text(),e35.text(),e36.text(),
				e37.text(),e38.text(),e39.text(),e40.text(),e41.text(),e42.text(),e44.text(),e45.text(),e46.text()))

			popup.setLayout(form)
			popup.setWindowTitle("View / Edit Details")
			popup.setWindowModality(Qt.ApplicationModal)
			popup.exec_()

		if action == depart:
			wb = load_workbook(excel_file_path)
			ws = wb.active
			if str(ws['O'+row_no].value) == "none":

				popup = QDialog()
				form = QFormLayout()

				ed1 = QDateEdit()
				ed1.setDate(QDate().currentDate())
				ed1.setCalendarPopup(True)
				ed1.setFont(QFont(font_style,font_size))
				form.addRow("DATE DEPARTURE",ed1)
				ed2 = QLineEdit()
				ed2.setFont(QFont(font_style,font_size))
				form.addRow("REMARKS",ed2)
				depart_btn = QPushButton("DONE")
				depart_btn.setFont(QFont("Courier",font_size))	
				form.addRow(depart_btn)
				depart_btn.clicked.connect(lambda:depart_btn_clkd(entry_no,row_no,str(ed1.date().toPyDate()),ed2.text(),popup))

				popup.setLayout(form)
				popup.setWindowTitle("Departure")
				popup.setWindowModality(Qt.ApplicationModal)
				popup.exec_()

			else:
				QMessageBox.information(tab1, "Error", "DEPARTURE ALREADY UPDATED")

		if action == delete:
			data = tablib.Dataset()
			data.headers = (entry_val)

			wb = load_workbook(excel_file_path)
			ws = wb.active
			row_cnt = ws.max_row
			column_cnt = ws.max_column
			for index,row in enumerate(ws.iter_rows()):
				if index>0 and index<row_cnt and str(row[0].value) != str(entry_no):
					temp = []
					for i in range(column_cnt):
						temp.append(row[i].value)
					data.append(temp)

			with open(excel_file_path, 'wb') as f:
				f.write(data.xlsx)

			QMessageBox.information(tab1, "Success", "ENTRY NO. "+str(entry_no)+" DELETED SUCCESSFULLY")


	def update_btn_clkd(trow,a1,a2,a3,a4,a5,a6,a7,a8,a9,a10,a11,a12,a13,a14,a15):
		wb = load_workbook(excel_file_path)
		ws = wb.active
		ws['A'+str(trow)] = str(a15)
		ws['C'+str(trow)] = str(a1)
		ws['D'+str(trow)] = str(a2)
		ws['E'+str(trow)] = str(a3)
		ws['F'+str(trow)] = str(a4)
		ws['G'+str(trow)] = str(a5)
		ws['H'+str(trow)] = str(a6)
		ws['I'+str(trow)] = str(a7)
		ws['J'+str(trow)] = str(a8)
		ws['K'+str(trow)] = str(a9)
		ws['L'+str(trow)] = str(a10)
		ws['M'+str(trow)] = str(a11)
		ws['N'+str(trow)] = str(a12)
		ws['P'+str(trow)] = str(a13)
		ws['Q'+str(trow)] = str(a14)
		wb.save(excel_file_path)
		QMessageBox.information(tab1, "Success", "DETAILS UPDATED SUCCESSFULLY")

	def depart_btn_clkd(entry_no,trow,a1,a2,popup):
		row_no = str(trow)
		wb = load_workbook(excel_file_path)
		ws = wb.active
		ws['O'+str(trow)] = str(a1)
		ws['P'+str(trow)] = str(a2)
		wb.save(excel_file_path)

		date_arr = str(ws['B'+row_no].value)
		name = str(ws['C'+row_no].value)
		age = str(ws['D'+row_no].value)
		nationality = str(ws['E'+row_no].value)
		addr = str(ws['F'+row_no].value)
		arrived_frm = str(ws['G'+row_no].value)
		addr_to = str(ws['H'+row_no].value)
		unique_id = str(ws['I'+row_no].value)
		purpose = str(ws['J'+row_no].value)
		occupation = str(ws['K'+row_no].value)
		person_cnt = str(ws['L'+row_no].value)
		date_dept = str(ws['O'+row_no].value)
		remarks = str(ws['P'+row_no].value)
		room_no = str(ws['M'+row_no].value)
		price = str(ws['N'+row_no].value)
		contact_no = str(ws['Q'+row_no].value)

		customer_det = [name,date_arr,date_dept,room_no,person_cnt,price,contact_no]

		html = render_template("receipt.html",customer_det=customer_det)
		print_pdf(html, receipt_path+str(entry_no)+".pdf")

		QMessageBox.information(tab1, "Success", "DEPARTURE TIME AND REMARKS UPDATED SUCCESSFULLY \nBILL HAS BEEN GENERATED")	
		popup.close()
	
	def enter_btn_submit():
		now = datetime.datetime.now()

		#now.strftime("%Y-%m-%d %H:%M")
		temp = [e14.text(),str(e13.date().toPyDate()),e0.text(),e1.text(),e2.text(),e3.text(),e4.text(),e5.text(),e6.text(),e7.text(),e8.text(),e9.text(),e10.text(),e11.text(),e12.text()]
		if "" not in temp:
			temp += ["none","none"]

			wb = load_workbook(excel_file_path)
			ws = wb.active
			row_cnt = ws.max_row
			column_cnt = ws.max_column
			ws['A'+str(row_cnt+1)] = str(temp[0])
			ws['B'+str(row_cnt+1)] = str(temp[1])
			ws['C'+str(row_cnt+1)] = str(temp[2])
			ws['D'+str(row_cnt+1)] = str(temp[3])
			ws['E'+str(row_cnt+1)] = str(temp[4])
			ws['F'+str(row_cnt+1)] = str(temp[5])
			ws['G'+str(row_cnt+1)] = str(temp[6])
			ws['H'+str(row_cnt+1)] = str(temp[7])
			ws['I'+str(row_cnt+1)] = str(temp[8])
			ws['J'+str(row_cnt+1)] = str(temp[9])
			ws['K'+str(row_cnt+1)] = str(temp[10])
			ws['L'+str(row_cnt+1)] = str(temp[11])
			ws['M'+str(row_cnt+1)] = str(temp[12])
			ws['N'+str(row_cnt+1)] = str(temp[13])
			ws['O'+str(row_cnt+1)] = "none"
			ws['P'+str(row_cnt+1)] = "none"
			ws['Q'+str(row_cnt+1)] = str(temp[14])
			wb.save(excel_file_path)

			QMessageBox.information(tab1, "Success", "DETAILS ENTERED SUCCESSFULLY")
			e0.setText("")
			e1.setText("")
			e2.setText("")
			e3.setText("")
			e4.setText("")
			e5.setText("")
			e6.setText("")
			e7.setText("")
			e8.setText("")
			e9.setText("")
			e10.setText("")
			e11.setText("")
			e12.setText("")
			e13.setDate(QDate().currentDate())
			e14.setText("")

		else:
			QMessageBox.warning(tab1, "Alert", "PLEASE ENTER ALL DETAILS")


	def amtcalc_btn_clkd(frm_dte,to_dte):
		tot_amt = 0
		wb = load_workbook(excel_file_path)
		ws = wb.active
		row_cnt = ws.max_row
		for index,row in enumerate(ws.iter_rows()):
			if index>0 and index<row_cnt:
				temp1 = parser.parse(row[1].value.split(" ")[0]).date()
				if(frm_dte <= temp1 and temp1 <= to_dte) and (str(row[14].value) != "none"):
					temp2 = parser.parse(row[14].value.split(" ")[0]).date()
					tdiff = int((temp2 - temp1).days)
					if tdiff == 0: tdiff=1
					tot_amt += float(row[13].value) * tdiff
		
		if tot_amt == 0:
			QMessageBox.information(tab1, "Sorry", "NO ENTRIES FOUND")
			el33.setText("TOTAL AMOUNT : N.A.")
		else:
			el33.setText("TOTAL AMOUNT : INR "+str(tot_amt)+" /-")

	def search_btn_clk():
		global tentryno, trowno, tpcnt, tname, tdatedeparture, tdatearr, table_data, troom
		visitor_name = str(e20.text())
		from_date = e21.date().toPyDate()
		to_date = e22.date().toPyDate()

		tentryno=[]
		tdatearr=[]
		tname=[]
		tdatedeparture=[]
		tpcnt=[]
		troom=[]
		trowno=[]

		wb = load_workbook(excel_file_path)
		ws = wb.active
		row_cnt = ws.max_row
		column_cnt = ws.max_column
		for index,row in enumerate(ws.iter_rows()):
			if index>0 and index<row_cnt:
				temp = parser.parse(row[1].value.split(" ")[0]).date()
				if(from_date <= temp and temp <= to_date):
					if (e23.isChecked()==True and str(row[14].value) == "none") or (e24.isChecked()==True and str(row[14].value) != "none") or(e25.isChecked()==True):
						if visitor_name == "":
							tentryno.append(str(row[0].value))
							tdatearr.append(str(row[1].value))
							tname.append(str(row[2].value))
							tpcnt.append(str(row[11].value))
							tdatedeparture.append(str(row[14].value))
							troom.append(str(row[12].value))
							trowno.append(str(index+1))
						elif str(row[2].value) == visitor_name:
							tentryno.append(str(row[0].value))
							tdatearr.append(str(row[1].value))
							tname.append(str(row[2].value))
							tpcnt.append(str(row[11].value))
							tdatedeparture.append(str(row[14].value))
							troom.append(str(row[12].value))
							trowno.append(str(index+1))

		if len(tentryno) != 0:
			table_data = {0:tentryno,1:tname,2:troom,3:tpcnt,4:tdatearr,5:tdatedeparture}
			table = MyTable(table_data, 0, 6)
			table.cellClicked.connect(cellClick)
			scroll.setWidget(table)
		else:
			QMessageBox.information(tab1, "Sorry", "NO ENTRIES FOUND")

	def render_template(template_file, **kwargs):
		return env.get_template(template_file).render(**kwargs)

	def print_pdf(html, destination):
		web = QWebView()
		web.setHtml(html)

		printer = QPrinter()
		printer.setPageSize(QPrinter.A4)
		printer.setOutputFormat(QPrinter.PdfFormat)
		printer.setOutputFileName(destination)
		web.print_(printer)

	env = Environment(loader=PackageLoader("vijay_niwas", pdf_template_path))

	app = QApplication(args)

	tabs = QTabWidget()

	tab1 = QWidget()
	tab2 = QWidget()
	tab3 = QWidget()

	Layout1 = QFormLayout()

	e14 = QLineEdit()
	e14.setValidator(QIntValidator())
	e14.setFont(QFont(font_style,font_size))
	Layout1.addRow("ENTRY NO",e14)

	e0 = QLineEdit()
	e0.setFont(QFont(font_style,font_size))
	e0.setFocus()
	Layout1.addRow("NAME OF VISITOR",e0)

	e12 = QLineEdit()
	e12.setValidator(QIntValidator())
	e12.setFont(QFont(font_style,font_size))
	Layout1.addRow("CONTACT NO",e12)

	e1 = QLineEdit()
	e1.setValidator(QIntValidator())
	e1.setMaxLength(3)
	e1.setFont(QFont(font_style,font_size))
	Layout1.addRow("AGE",e1)

	e2 = QLineEdit("Indian")
	e2.setFont(QFont(font_style,font_size))
	Layout1.addRow("NATIONALITY",e2)

	e3 = QLineEdit()
	e3.setFont(QFont(font_style,font_size))	
	Layout1.addRow("ADDRESS",e3)

	e4 = QLineEdit()
	e4.setFont(QFont(font_style,font_size))	
	Layout1.addRow("ARRIVED FROM",e4)

	e5 = QLineEdit()
	e5.setFont(QFont(font_style,font_size))
	Layout1.addRow("ADDRESS TO WHICH PRECEDING",e5)

	e6 = QLineEdit()
	e6.setFont(QFont(font_style,font_size))
	Layout1.addRow("UNIQUE IDENTIFICATION",e6)

	e7 = QLineEdit()
	e7.setFont(QFont(font_style,font_size))	
	Layout1.addRow("PURPOSE OF VISIT",e7)

	e8 = QLineEdit()
	e8.setFont(QFont(font_style,font_size))	
	Layout1.addRow("OCCUPATION",e8)

	e9 = QLineEdit()
	e9.setValidator(QIntValidator())
	e9.setMaxLength(3)
	e9.setFont(QFont(font_style,font_size))	
	Layout1.addRow("TOTAL NO. OF PERSONS",e9)

	e10 = QLineEdit()
	e10.setValidator(QIntValidator())
	e10.setFont(QFont(font_style,font_size))	
	Layout1.addRow("ROOM NO",e10)

	e11 = QLineEdit()
	e11.setValidator(QIntValidator())
	e11.setFont(QFont(font_style,font_size))	
	Layout1.addRow("PRICE",e11)

	e13 = QDateEdit()
	e13.setDate(QDate().currentDate())
	e13.setCalendarPopup(True)
	Layout1.addRow("DATE ARRIVAL",e13)

	enter_btn = QPushButton("SUBMIT")
	enter_btn.setFont(QFont("Courier",font_size))	
	Layout1.addRow(enter_btn)
	enter_btn.clicked.connect(lambda:enter_btn_submit())

	tab1.setLayout(Layout1)
	#tab1.connect(enter_btn, SIGNAL("clicked()"),enter_btn_submit())

	
	Layout2 = QVBoxLayout()
	Layout21 = QFormLayout()
	Layout22 = QGridLayout()

	e20 = QLineEdit()
	e20.setFont(QFont(font_style,font_size))
	Layout21.addRow("VISITOR NAME",e20)

	e21 = QDateEdit()
	e21.setDate(QDate().currentDate())
	e21.setCalendarPopup(True)
	Layout21.addRow("FROM DATE",e21)

	e22 = QDateEdit()
	e22.setDate(QDate().currentDate())
	e22.setCalendarPopup(True)
	Layout21.addRow("TO DATE",e22)

	e23 = QRadioButton("CURRENT")
	e24 = QRadioButton("DEPARTED")
	e25 = QRadioButton("BOTH")
	e23.setChecked(True)
	Layout22.addWidget(e23,0,1)
	Layout22.addWidget(e24,0,2)
	Layout22.addWidget(e25,0,3)

	search_btn = QPushButton("SEARCH")
	search_btn.setFont(QFont("Courier",font_size))	
	Layout21.addRow(search_btn)
	search_btn.clicked.connect(lambda:search_btn_clk())

	Layout2.addLayout(Layout22)
	Layout2.addLayout(Layout21)
	
	scroll = QScrollArea()
	scroll.setWidgetResizable(True)
	Layout2.addWidget(scroll)

	tab2.setLayout(Layout2)

	Layout3 = QGridLayout()
	Layout31 = QFormLayout()

	el31 = QDateEdit()
	el31.setDate(QDate().currentDate())
	el31.setCalendarPopup(True)
	Layout31.addRow("FROM DATE",el31)

	el32 = QDateEdit()
	el32.setDate(QDate().currentDate())
	el32.setCalendarPopup(True)
	Layout31.addRow("TO DATE",el32)

	amtcalc_btn = QPushButton("CALCULATE")
	amtcalc_btn.setFont(QFont("Courier",font_size))	
	Layout31.addRow(amtcalc_btn)
	amtcalc_btn.clicked.connect(lambda:amtcalc_btn_clkd(el31.date().toPyDate(),el32.date().toPyDate()))

	dummy_label = QLabel("")
	el33 = QLabel("TOTAL AMOUNT : N.A.")
	el33.setFont(QFont("Courier",font_size))	
	el33.setAlignment(Qt.AlignCenter)
	Layout31.addWidget(dummy_label)
	Layout31.addWidget(el33)

	Layout3.addLayout(Layout31,0,0)

	tab3.setLayout(Layout3)

	tabs.resize(550,550)

	tabs.addTab(tab1,"New Entry")
	tabs.addTab(tab2,"Search Entries")
	tabs.addTab(tab3,"Amount Calculator")

	tabs.setWindowTitle("Vijay Niwas Lodge")

	tabs.show()

	app_icon = QIcon()
	app_icon.addFile('database/icon/16x16.png', QSize(16,16))
	app_icon.addFile('database/icon/24x24.png', QSize(24,24))
	app_icon.addFile('database/icon/32x32.png', QSize(32,32))
	app_icon.addFile('database/icon/48x48.png', QSize(48,48))
	app_icon.addFile('database/icon/256x256.png', QSize(256,256))
	app.setWindowIcon(app_icon)

	sys.exit(app.exec_())

if __name__=="__main__":
    main(sys.argv)